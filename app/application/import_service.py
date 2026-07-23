from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import UploadFile


class ImportParseError(Exception):
    def __init__(self, message: str) -> None:
        self.message = message
        super().__init__(message)


@dataclass(frozen=True)
class ImportRow:
    row_number: int
    values: dict[str, str]


SCHOOL_COLUMNS = {
    "name": ["nom", "nom etablissement", "nom établissement", "name"],
    "email": ["email", "e-mail", "adresse email", "adresse e-mail"],
    "address": ["adresse", "address"],
    "city": ["ville", "city"],
    "postalCode": ["code postal", "postalcode", "postal code"],
    "phone": ["telephone", "téléphone", "phone"],
    "directorName": ["directeur", "directrice", "nom directeur", "director"],
    "logoUrl": ["logo", "logo url", "url logo"],
}

PROFESSOR_COLUMNS = {
    "firstName": ["prenom", "prénom", "first name", "firstname"],
    "lastName": ["nom", "last name", "lastname"],
    "email": ["email", "e-mail", "adresse email", "adresse e-mail"],
    "phone": ["telephone", "téléphone", "phone"],
    "dateOfBirth": ["date naissance", "date de naissance", "birth date", "dateofbirth"],
}


def template_csv(kind: str) -> str:
    if kind == "schools":
        headers = [
            "Nom",
            "Email",
            "Adresse",
            "Ville",
            "Code postal",
            "Téléphone",
            "Directeur",
            "Logo URL",
        ]
        sample = [
            "École Victor Hugo",
            "contact@victor-hugo.fr",
            "12 rue de la Paix",
            "Paris",
            "75001",
            "+33123456789",
            "Marie Dupont",
            "",
        ]
    else:
        headers = ["Prénom", "Nom", "Email", "Téléphone", "Date de naissance"]
        sample = ["Marie", "Dupont", "marie.dupont@ecole.fr", "+33123456789", "1985-04-12"]
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample)
    return output.getvalue()


async def parse_import_file(file: UploadFile, *, kind: str) -> list[ImportRow]:
    content = await file.read()
    suffix = Path(file.filename or "").suffix.lower()
    aliases = SCHOOL_COLUMNS if kind == "schools" else PROFESSOR_COLUMNS
    if suffix == ".csv":
        return _parse_csv(content, aliases)
    if suffix == ".xlsx":
        return _parse_xlsx(content, aliases)
    raise ImportParseError("Format non supporté. Utilisez un fichier CSV ou XLSX.")


def _parse_csv(content: bytes, aliases: dict[str, list[str]]) -> list[ImportRow]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ImportParseError("Le fichier ne contient pas d'en-têtes.")
    return _rows_from_dicts(reader, aliases)


def _parse_xlsx(content: bytes, aliases: dict[str, list[str]]) -> list[ImportRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportParseError("Le support XLSX n'est pas installé.") from exc
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ImportParseError("Le fichier ne contient pas d'en-têtes.")
    headers = [str(value or "") for value in rows[0]]
    dict_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        dict_rows.append({headers[index]: value for index, value in enumerate(row)})
    return _rows_from_dicts(dict_rows, aliases)


def _rows_from_dicts(rows: Any, aliases: dict[str, list[str]]) -> list[ImportRow]:
    parsed: list[ImportRow] = []
    for index, row in enumerate(rows, start=2):
        values = _normalize_row(row, aliases)
        if any(values.values()):
            parsed.append(ImportRow(row_number=index, values=values))
    if not parsed:
        raise ImportParseError("Le fichier ne contient aucune ligne à importer.")
    return parsed


def _normalize_row(row: dict[str, Any], aliases: dict[str, list[str]]) -> dict[str, str]:
    normalized_headers = {_normalize_header(key): value for key, value in row.items()}
    values: dict[str, str] = {}
    for target, names in aliases.items():
        value = ""
        for name in names:
            raw = normalized_headers.get(_normalize_header(name))
            if raw is not None:
                value = str(raw).strip()
                break
        values[target] = value
    return values


def _normalize_header(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("_", " ")
        .replace("-", " ")
    )
